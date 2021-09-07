import errno
import json
import os
import sys
import threading
import openpyxl
import pandas
import requests
import wx
import base64
import datetime
import time
import locale
import re


class BulkMailer(wx.Frame):
    def __init__(self, parent, fid):
        wx.Frame.__init__(self, parent, fid, "Bulk Mailer", size=(580, 480))
        self.panel = wx.Panel(self)
        try:
            image_file = os.path.join(util_dir, 'mibs.png')
            image_file2 = os.path.join(util_dir, 'rise.png')
            bmp1 = wx.Image(image_file, wx.BITMAP_TYPE_ANY).ConvertToBitmap()
            self.bitmap1 = wx.StaticBitmap(self.panel, -1, bmp1, (120, 10))
            bmp2 = wx.Image(image_file2, wx.BITMAP_TYPE_ANY).ConvertToBitmap()
            self.bitmap2 = wx.StaticBitmap(self.panel, -1, bmp2, (460, 415))
        except:
            pass

        self.win1 = wx.Panel(self.panel, -1, pos=(380, 145), size=(160, 130))
        self.textwin1 = wx.Panel(self.panel, -1, pos=(380, 145), size=(160, 35))
        self.midwin = wx.Panel(self.panel, -1, pos=(410, 197), size=(100, 60), style=wx.BORDER_THEME)
        self.wintext = wx.StaticText(self.panel, -1, "E-mail Counter", pos=(415, 153))
        self.counttext = wx.StaticText(self.panel, -1, "0000", pos=(433, 212))
        self.win1.SetBackgroundColour("#5ba6e3")
        self.textwin1.SetBackgroundColour("#e31837")
        self.titletext = wx.StaticText(self.panel, -1, "Bulk E-Mails", pos=(240, 115))

        self.subtext = wx.TextCtrl(self.panel, pos=(160, 145), size=(200, 20))
        self.sublabel = wx.StaticText(self.panel, -1, "Subject*:", (35, 147))

        self.campNtext = wx.TextCtrl(self.panel, pos=(160, 167), size=(200, 20))
        self.campNlabel = wx.StaticText(self.panel, -1, "Campaign Name*:", (35, 169))

        self.tbodytext = wx.TextCtrl(self.panel, pos=(160, 189), size=(200, 20))
        self.tbodylabel = wx.StaticText(self.panel, -1, "Text Body*:", (35, 191))

        self.listNtext = wx.TextCtrl(self.panel, pos=(160, 211), size=(200, 20))
        self.listNlabel = wx.StaticText(self.panel, -1, "List Name*:", (35, 213))

        self.filebtn = wx.Button(self.panel, -1, label="Browse", pos=(159, 233))
        self.filelabel = wx.StaticText(self.panel, -1, "Html File*:", (35, 235))
        self.file2label = wx.StaticText(self.panel, -1, "", (240, 235))

        self.attachfilebtn = wx.Button(self.panel, -1, label="Browse", pos=(159, 257))
        self.attachfilelabel = wx.StaticText(self.panel, -1, "Attachment:", (35, 259))
        self.attachfile2label = wx.StaticText(self.panel, -1, "", (240, 259))

        self.start_button = wx.Button(self.panel, -1, label="START", pos=(110, 360), size=(100, 30))
        self.stop_button = wx.Button(self.panel, -1, label="STOP", pos=(350, 360), size=(100, 30))
        self.Bind(wx.EVT_BUTTON, self.start_thread, self.start_button)
        self.Bind(wx.EVT_BUTTON, self.stop_f, self.stop_button)
        self.Bind(wx.EVT_BUTTON, self.open_fl, self.filebtn)
        self.Bind(wx.EVT_BUTTON, self.open_attachment, self.attachfilebtn)
        self.copyrightlabel = wx.StaticText(self.panel, -1,
                                            "Copyright © Mahindra Integrated Business Solutions Pvt. Ltd.", (10, 420))

        self.messagelabel = wx.StaticText(self.panel, -1,
                                          "Note: Excel file should contain only one E-mail per row.", (130, 300))
        self.messagelabel.Wrap(300)
        self.messagelabel.SetForegroundColour("blue")

        font = wx.Font(10, wx.FONTFAMILY_ROMAN, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_BOLD)

        font2 = wx.Font(12, wx.FONTFAMILY_ROMAN, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_BOLD)
        font5 = wx.Font(12, wx.FONTFAMILY_ROMAN, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_BOLD)
        font5.SetUnderlined(True)
        font3 = wx.Font(20, wx.FONTFAMILY_ROMAN, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_BOLD)
        font4 = wx.Font(11, wx.FONTFAMILY_ROMAN, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_BOLD)
        self.messagelabel.SetFont(font4)
        self.campNlabel.SetFont(font4)
        self.tbodylabel.SetFont(font4)
        self.listNlabel.SetFont(font4)
        self.sublabel.SetFont(font4)
        self.filelabel.SetFont(font4)
        self.attachfilelabel.SetFont(font4)

        self.counttext.SetFont(font3)
        self.start_button.SetBackgroundColour("#e31837")
        self.stop_button.SetBackgroundColour("#e31837")
        self.start_button.SetFont(font)
        self.titletext.SetFont(font5)
        self.wintext.SetFont(font2)
        self.wintext.SetBackgroundColour("#e31837")
        self.campNtext.SetBackgroundColour("#5ba6e3")
        self.tbodytext.SetBackgroundColour("#5ba6e3")
        self.listNtext.SetBackgroundColour("#5ba6e3")
        self.subtext.SetBackgroundColour("#5ba6e3")
        self.start_button.SetForegroundColour("#FAFAFA")
        self.stop_button.SetForegroundColour("#FAFAFA")

    def open_fl(self, event):
        try:
            self.htmlfile = wx.FileDialog(self.panel, "Open", "", "", "Html files (*.html)|*.html",
                                          wx.FD_OPEN | wx.FD_FILE_MUST_EXIST, pos=(160, 253))
            self.htmlfile.ShowModal()
            if self.htmlfile.GetPath():
                self.file2label.SetLabel(str(os.path.basename(self.htmlfile.GetPath()))[:16] + "...")
            file = open(self.htmlfile.GetPath(), "r").read()
            self.base64_message = self.get_base64Encode(file)
        except:
            pass

    def open_attachment(self, event):
        try:
            self.attachfile = wx.FileDialog(self.panel, "Open", "", "", "Text files (*.txt)|*.txt",
                                            wx.FD_OPEN | wx.FD_FILE_MUST_EXIST, pos=(160, 253))
            self.attachfile.ShowModal()
            if self.attachfile.GetPath():
                self.attachfile2label.SetLabel(str(os.path.basename(self.attachfile.GetPath()))[:16] + "...")
            file = open(self.attachfile.GetPath(), "r").read()
            self.base64_attachment = self.get_base64Encode(file)
        except:
            pass

    def get_base64Encode(self, data):
        message_bytes = data.encode('ascii')
        base64_bytes = base64.b64encode(message_bytes)
        return base64_bytes.decode('ascii')

    def get_base64Decode(self, data):
        base64_bytes = data.encode('ascii')
        message_bytes = base64.b64decode(base64_bytes)
        message = message_bytes.decode('ascii')
        return message

    def stop_f(self, event):
        try:
            if self.responsedata:
                self.create_excel(self.responsedata, self.Out_Excel)
                self.responsedata = []
                self.create_mis(
                    [self.campname, self.listname, self.todaydate, self.totalcount, self.successcount, self.failcount],
                    Mis_Excel)
        except:
            pass
        sys.exit()

    def start_thread(self, event):
        self.todaydate = get_datetime()
        self.campname = self.campNtext.GetValue()
        self.bodytext = self.tbodytext.GetValue()
        self.listname = self.listNtext.GetValue()
        self.subject = self.subtext.GetValue()

        try:
            self.config_data = json.load(open(CONFIG_PATH))
        except FileNotFoundError:
            self.messagelabel.SetLabel("Unable to Proceed. Config File not Found.")
            self.messagelabel.SetForegroundColour("red")
            return

        if not (self.campname and self.listname and self.subject and self.bodytext):
            self.messagelabel.SetLabel("Please fill all the Fields")
            self.messagelabel.SetForegroundColour("red")
            return
        try:
            self.htmlpath = self.htmlfile.GetPath()
        except AttributeError:
            self.messagelabel.SetLabel("Please select a Html File.")
            self.messagelabel.SetForegroundColour("red")
            return
        try:
            self.attpath = self.attachfile.GetPath()
        except AttributeError:
            self.attpath = False

        if self.campname and self.listname and self.subject and self.htmlfile.GetPath():
            self.t1 = threading.Thread(target=self.read_data)
            self.t1.setDaemon(True)
            self.t1.start()

    def read_data(self):
        t1 = int(time.time())
        self.start_button.Disable()
        self.messagelabel.SetLabel("")
        self.responsedata = []
        self.cyclecount = 0
        self.totalcount = 0
        self.successcount = 0
        self.failcount = 0
        self.Out_Excel = os.path.join(output_dir, f"{self.campname}_{get_date()}.xlsx")
        filename = os.path.join(input_dir, "Inputfile.xlsx")
        df = pandas.read_excel(filename)
        n = 0
        regex = '^[a-z0-9]+[\._]?[a-z0-9]+[@]\w+[.]\w{2,3}$'
        data = []
        for x in df.values:
            if not (re.search(regex, x[2])):
                if len(re.findall("@", x[2])) > 1:
                    self.messagelabel.SetLabel(f"Unable to Proceed. Only one E-mail per row is allowed: {x[2]}")
                else:
                    self.messagelabel.SetLabel(f"Unable to Proceed. Invalid email found: {x[2]}")
                self.messagelabel.Wrap(300)
                self.messagelabel.SetForegroundColour("red")
                self.start_button.Enable()
                return
            data.append([x[0], x[1], x[2], x[3]])

        chunkrange = 20
        chunks = [data[x:x + chunkrange] for x in range(0, len(data), chunkrange)]
        for x in chunks:
            if n >= 49:
                n = 0
                self.messagelabel.SetLabel(
                    f"Reached the mail limit of 980 mails, System will halt for 3 minutes. System will resume again at {(datetime.datetime.now() + datetime.timedelta(minutes=3)).strftime('%H:%M')}")
                self.messagelabel.Wrap(300)
                self.messagelabel.SetForegroundColour("red")
                self.create_excel(self.responsedata, self.Out_Excel)

                self.responsedata = []
                self.cyclecount = 0
                time.sleep(180)
                self.messagelabel.SetForegroundColour("green")
                self.messagelabel.SetLabel("Resuming now.")
                time.sleep(2)
                self.messagelabel.SetLabel("")
                self.messagelabel.SetForegroundColour("red")

            n += 1
            self.send_mail(x)
        if self.responsedata:
            self.create_excel(self.responsedata, self.Out_Excel)
            self.responsedata = []
        self.create_mis(
            [self.campname, self.listname, self.todaydate, self.totalcount, self.successcount, self.failcount],
            Mis_Excel)
        self.messagelabel.SetLabel("All Mails Sent.")
        self.messagelabel.SetForegroundColour("Green")
        print("Total time taken", int(time.time()) - t1)

    def send_mail(self, rec):
        tomail = []
        for x in rec:
            tomail.append(x[2])

        url = 'http://api.emsender.in/campaign_api/campaign/format/json'
        headers = {
            "Content-Type": "application/json",
            "Accept-Encoding": "identity"
        }
        body = {
            "username": self.get_base64Decode(self.config_data["username"]),
            "password": self.get_base64Decode(self.config_data["password"]),
            "from": self.config_data["from"],
            "fromName": self.config_data["fromName"],
            "to": tomail,
            "subject": self.subject,
            "replyTo": self.config_data["replyTo"],
            "campaignName": self.campname,
            "htmlBody": self.base64_message,
            "textBody": self.bodytext,
            "listName": self.listname,

        }
        if self.attpath:
            body["attachments"] = [{"name": os.path.basename(self.attachfile.GetPath()),
                                    "content": self.base64_attachment,
                                    "contentType": "text/plain"}]
        res = requests.post(url, headers=headers, data=json.dumps(body), verify=False)

        # res = requests.post(url, headers=headers, data=json.dumps(body))
        self.responselist = []
        self.responselist.append(self.listname)
        res_data = json.loads(res.text)
        if str(res_data["code"]) == "200":
            self.responselist.append("Success")
            self.successcount += len(rec)
        else:
            self.responselist.append("Fail")
            self.failcount += 1
        self.responselist.append(str(res_data["message"]).split(":-")[1])
        self.responselist.append(res_data["code"])
        self.responselist.append(res_data["message"])
        self.responselist.append(datetime_from_utc_to_local(str(res.headers["Date"])))

        for item in rec:
            print(item)
            self.cyclecount += 1
            self.totalcount += 1
            self.counttext.SetLabel(str(self.totalcount).zfill(4))
            item.extend(self.responselist)
            self.responsedata.append(item)

    def create_excel(self, allrecord, fpath):
        if not os.path.isfile(fpath):
            book = openpyxl.Workbook()
            sheet = book.active
            sheet.append(["Sl. No.", "Company Name", "Email", "Campaign", "Campaign List", "Result", "Campaign Id",
                          "Server response Code", "Server Response Message", "Date and time"])
        else:
            book = openpyxl.load_workbook(fpath)
            sheet = book.active
        for record in allrecord:
            sheet.append(record)
        book.save(fpath)

    def create_mis(self, allrecord, fpath):
        if not os.path.isfile(fpath):
            book = openpyxl.Workbook()
            sheet = book.active
            sheet.append(["Campaign Name", "List Name", "Date", "Total Count", "Success Count", "Fail Count"])
        else:
            book = openpyxl.load_workbook(fpath)
            sheet = book.active
        sheet.append(allrecord)
        book.save(fpath)


def get_date():
    dt = datetime.datetime.now()
    date = dt.strftime("%d%m%Y")
    return str(date)


def get_datetime():
    dt = datetime.datetime.now()
    date = dt.strftime("%a, %d %b %Y %H:%M:%S")
    return str(date)


def datetime_from_utc_to_local(utc_datetime):
    locale.setlocale(locale.LC_ALL, "en_US.UTF-8")
    now_timestamp = time.time()
    offset = datetime.datetime.fromtimestamp(now_timestamp) - datetime.datetime.utcfromtimestamp(now_timestamp)
    return str((datetime.datetime.strptime(utc_datetime, "%a, %d %b %Y %H:%M:%S %Z") + offset).strftime(
        "%a, %d %b %Y %H:%M:%S %Z")) + "IST"


def make_dir(*paths):
    for pt in paths:
        if not (os.path.isdir(pt)):
            try:
                os.makedirs(pt, mode=0o777, exist_ok=True)
            except OSError as exception:
                if exception.errno != errno.EEXIST:
                    raise


if __name__ == "__main__":
    input_dir = os.path.join(os.getcwd(), "Utility Files", "Input Files")
    config_dir = os.path.join(os.getcwd(), "Utility Files", "Config Files")
    util_dir = os.path.join(config_dir, "util")
    output_dir = os.path.join(os.getcwd(), "Utility Files", "Output Files")
    make_dir(input_dir, output_dir, config_dir, util_dir)

    CONFIG_PATH = os.path.join(config_dir, "config.json")

    Mis_Excel = os.path.join(output_dir, "MIS.xlsx")

    app = wx.App()
    frame = BulkMailer(parent=None, fid=-1)
    frame.Show()
    app.MainLoop()
